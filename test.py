import os
from docx import Document
import openpyxl
from openpyxl.styles import Font
import re
from docx.table import Table
from docx.text.paragraph import Paragraph

def iter_block_items(parent):
    """
    Generate a reference to each paragraph and table child within *parent*,
    in document order. Each returned value is an instance of either Table or
    Paragraph. *parent* would most commonly be a reference to a main
    Document object, but also works for a _Cell object, which itself can
    contain paragraphs and tables.
    """
    parent_elm = parent.element.body

    for child in parent_elm.iterchildren():
        if child.tag.endswith('p'):
            yield Paragraph(child, parent)
        elif child.tag.endswith('tbl'):
            yield Table(child, parent)

def extract_tables_and_headings(docx_path):
    doc = Document(docx_path)
    tables_and_headings = []
    blocks = list(iter_block_items(doc))

    # Regular expression to match the required patterns
    pattern = re.compile(r'^(Table \d+:|Figure \d+:|Table :|Figure :)')

    last_heading = None

    for block in blocks:
        if isinstance(block, Paragraph):
            # Check if the paragraph matches any of the specified patterns
            if pattern.match(block.text.strip()):
                last_heading = block.text.strip()
        elif isinstance(block, Table):
            # If a table is found, associate it with the last heading
            if last_heading:
                tables_and_headings.append((last_heading, block))
                last_heading = None

    return tables_and_headings

def create_excel_from_directory(directory_path, excel_path):
    wb = openpyxl.Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Creating headers for summary sheet
    headers = ["Document Name", "Table Heading", "Link to Table"]
    summary_sheet.append(headers)
    for col in range(1, len(headers) + 1):
        summary_sheet.cell(row=1, column=col).font = Font(bold=True)

    for filename in os.listdir(directory_path):
        if filename.endswith('.docx'):
            docx_path = os.path.join(directory_path, filename)
            doc_name = os.path.splitext(filename)[0]
            tables_and_headings = extract_tables_and_headings(docx_path)
            
            for idx, (heading, table) in enumerate(tables_and_headings, start=1):
                sheet_name = f"{doc_name}_Table_{idx}"
                table_sheet = wb.create_sheet(title=sheet_name)
                
                # Write table data to the table_sheet
                for i, row in enumerate(table.rows, start=1):
                    for j, cell in enumerate(row.cells, start=1):
                        table_sheet.cell(row=i, column=j).value = cell.text

                # Adding data to the summary sheet
                summary_sheet.append([doc_name, heading, f"=HYPERLINK(\"#{sheet_name}!A1\", \"{sheet_name}\")"])
                link_cell = summary_sheet.cell(row=summary_sheet.max_row, column=3)
                # Set the font to blue and underlined
                link_cell.font = Font(color="0000FF", underline="single")

    wb.save(excel_path)

